import mysql.connector
import datetime as dt 
import openpyxl
import numpy
import os
import smtplib
import email.encoders
from configparser import ConfigParser
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# List of person type 1, type 2 and type 3
# bien toan cuc phai in hoa LIST_TYPE1
LIST_TYPE1 = ['P2', 'P3', 'P4', 'P5', 'P6','P1']
LIST_TYPE2 = ['PA', 'PB', 'PC', 'PD']
LIST_TYPE3 = ['DuongPT3']

class sql():
    def __init__(self):
        config = ConfigParser()
        config.read(os.getcwd() + "/config.txt")
        databaseHost = config.get("Database", "host")
        databaseUsername = config.get("Database", "user")
        databasePassword = config.get("Database", "password")
        databaseDB = config.get("Database", "database")
        query1 = config.get("Database", "query1")
        query2 = config.get("Database", "query2")
        #----------------------------------------------------
        self.username = config.get("Gmail", "Account")
        self.password = config.get("Gmail", "Password")
        self.msg = MIMEMultipart()
        self.msg['From'] = self.username
        self.msg['To'] = self.username
        self.msg['Subject'] = config.get("Gmail", "Subject")
        self.body = config.get("Gmail", "Body")
        self.filename = config.get("Gmail", "Filename")
        #-----------------------------------------------
        self.mydb = mysql.connector.connect(
        host= databaseHost,
        user=databaseUsername,
        passwd=databasePassword)
        self.cursor = self.mydb.cursor()
        self.cursor.execute("USE " + databaseDB)
        self.cursor.execute(query1)
        self.cursor.execute(query2)
        self.creat_and_desgin_excel()
        
    @staticmethod
    def creat_and_desgin_excel():
        # Create excel file
        wb = openpyxl.Workbook()
        sheet = wb.active    
        # Design                 
        sheet.merge_cells('A1:J1') 
        sheet['A1'] = 'LICH TRUC GIAM SAT ATTT' 
        sheet.merge_cells('A2:B2')
        sheet['A2'] = 'THOI GIAN: 30 NGAY'
        sheet.merge_cells('A3:A5') 
        sheet['A3'] = 'Ngay'
        sheet.merge_cells('B3:J3') 
        sheet['B3'] = 'Ca truc'
        sheet.merge_cells('B4:D4') 
        sheet['B4'] = 'Ca 1 (0:00 - 8:00)'
        sheet.merge_cells('E4:G4') 
        sheet['E4'] = 'Ca 2 (8:00 - 16:00)'
        sheet.merge_cells('H4:J4') 
        sheet['H4'] = 'Ca 3 (16:00 - 0:00)'
        sheet['B5'] = 'Tier 1' 
        sheet['C5'] = 'Tier 2'
        sheet['D5'] = 'Tier 3' 
        sheet['E5'] = 'Tier 1'        
        sheet['F5'] = 'Tier 2' 
        sheet['G5'] = 'Tier 3'
        sheet['H5'] = 'Tier 1' 
        sheet['I5'] = 'Tier 2'
        sheet['J5'] = 'Tier 3'
        wb.save('Schedule_Ex1_Update.xlsx')


    def schedule(self, day, LIST_TYPE1, LIST_TYPE2, LIST_TYPE3):
        wb = openpyxl.load_workbook('Schedule_Ex1_Update.xlsx')
        sheet = wb.active
        count = 0
        count1 = 0
        a = 0
        b = 0
        i = 0
        line = ''
        while i < int(day):
            for j in range (3):
                # 1 shift
                line1 = LIST_TYPE1[a] + " " + LIST_TYPE2[b] + " " + LIST_TYPE3[0] + " "
                
                # Write to mysql
                today = dt.date.today()
                name1 = str((LIST_TYPE1[a]))
                timestamp1 = (str(today + dt.timedelta(days = (i + 1))))
                shift1 = (str((j + 1)))
                # Avoid sql injection
                query3 = """INSERT INTO TASK_SCHEDULE(name, timestamp, shift) 
                            VALUES(%s, %s, %s);
                         """,  (name1, timestamp1, shift1)
                self.cursor.execute(*query3)           
                # Write to excel
                sheet.cell(row= i + 6, column = 1).value = timestamp1
                
                # 1 day equal 3 shifts 
                line += line1
                count +=1
                count1 +=1
                ''' if count == 6 need roll list person of type 1 1 time to avoid
                one person of list type 1 have slot 1 2 times in a row
                (truong hop 1 nguoi type 1 lam slot 1 2 lan lien tiep) '''
                if count == 6:
                    if count1 == 12:
                        LIST_TYPE1 = numpy.roll(LIST_TYPE1, -2)
                        a = 0
                        count = 0
                        count1 = 0
                    else:
                        LIST_TYPE1 = numpy.roll(LIST_TYPE1, -1)
                        a = 0
                        count = 0
                else:
                    # if count != 6 increasing a 1 unit
                    a = a + 1
            # khi type 2 dc goi het thì se goi lai tu dau
            if (b == 3):
                b = 0
            else:
                b = b + 1
            # Write to sql
            name2 = LIST_TYPE2[b]
            name3 = LIST_TYPE3[0]
            timestamp2 = str(today + dt.timedelta(days = (i + 1)))
            shift2 = '1, 2, 3'
            query4 = """INSERT INTO TASK_SCHEDULE(name, timestamp, shift) 
                        VALUES(%s, %s, %s);
                     """, (name2, timestamp2, shift2)
            query5 = """INSERT INTO TASK_SCHEDULE(name, timestamp, shift) 
                        VALUES(%s, %s, %s);
                     """, (name3, timestamp2, shift2)
            self.cursor.execute(*query4)
            self.cursor.execute(*query5)        
            self.mydb.commit()
            # Write to excel
            array = line.split(' ')
            for k in range (len(array)):
                sheet.cell(row = i + 6, column = k + 2).value =  array[k]
            i = i + 1
            line = ''
        wb.save('Schedule_Ex1_Update.xlsx')
        wb.close()


    def send_email(self):
        try:
            self.msg.attach(MIMEText(self.body, 'plain'))
            attachment = open(self.filename, "rb")
            part = MIMEBase('application', 'octet-stream')
            part.set_payload((attachment).read())
            email.encoders.encode_base64(part)
            part.add_header('Content-Disposition', "attachment; filename = %s" %self.filename)
            self.msg.attach(part)
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.username, self.password)
            text = self.msg.as_string()
            server.sendmail(self.username, 'hieutt35@fpt.com.vn', text)
            server.quit
        except Exception as e:
            print(str(e))
    

def main():
    sqlConnect = sql()
    sqlConnect.schedule(30, LIST_TYPE1, LIST_TYPE2, LIST_TYPE3)
    sqlConnect.send_email()

if __name__ == '__main__':
    main()
